from tkinter import *
#from PIL import ImageTk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pickle import dump
import smtplib
from email.message import EmailMessage
import ssl


window = Tk()
window.title("Attendence")
window.rowconfigure(0, weight=1)
window.columnconfigure(0, weight=1)
window.resizable(False,False)
window.geometry('1520x830')

Login = Frame(window)
page1 = Frame(window)

#window.configure(bg="#fff")
for frame in (Login, page1):
    frame.grid(row=0, column=0, sticky='nsew')

def log_out():
    txt_pass.delete(0,"end")
    Login.tkraise()

Login.tkraise()

#=======================================================

def forgot(userid):
    try:
        if userid =="" or userid==" ":
            messagebox.showerror("Error","Username required")
            return
        #af=open("attendence/adminlogin.dat","rb")
        #f=open("attendence/login.dat","rb")
        af=open("adminlogin.dat","rb")
        f=open("login.dat","rb")

        for r in af.readlines() :
            if userid in str(r) :
                key=str(r) 
        for r in f.readlines():
            if userid in str(r) :
                key=str(r)
        key=key[::-1]
        ps=""
        for i in key:
            if i==" ":
                break
            ps=ps+i
        ps=ps[3:]
        ps=ps[::-1]


        email_sender='attendence.login2022@gmail.com'
        email_password= ""
        email_receiver = 'anshulverma1405@gmail.com'

        subject = "Request to forgot passward"

        body=f'''
This mail is with the regards to the request to forget password 
made by a user of Attendence softwate .
the username and password for the same are given below .

Username : {userid}
password : {ps}

'''
        en=EmailMessage()

        en['From'] = email_sender
        en['To'] = email_receiver
        en['Subject']= subject
        en.set_content(body)
        context = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            smtp.login(email_sender, email_password)
            smtp.sendmail (email_sender, email_receiver, en.as_string())
        messagebox.showinfo("Password send",f"Password is been mailed to your admin at {email_receiver}")
    except:
        messagebox.showerror("Error","Unable to forgot password")
#=======================================================

def deletentry(userid):
    try:
        if userid ==" " :
            s = True
        else:
            s = False
        #af=open("attendence/adminlogin.dat","rb")
        #f=open("attendence/login.dat","rb")
        af=open("adminlogin.dat","rb")
        f=open("login.dat","rb")
        for r in af.readlines() :
            if userid in str(r) :
                s = "admin"
        for r in f.readlines():
            if userid in str(r) :
                s = "login"
        f.close()

        if s == "admin" :
            #f=open("attendence/adminlogin.dat","rb")
            f=open("adminlogin.dat","rb")
            data=f.readlines()
            f.close()
            f=open("adminlogin.dat","wb")
            for i in data:
                if userid in str(i):
                    continue
                dump(i,f) 
            f.close()
            messagebox.showinfo("Sucess","Entry deleted sucessfully")
    
        elif s == "login":
            #f=open("attendence/login.dat","rb") 
            f=open("login.dat","rb")
            data=f.readlines()
            f.close()
            f=open("login.dat","wb")
            for i in data:
                if userid in str(i):
                    continue
                dump(i,f)  
            f.close()
            messagebox.showinfo("Sucess","Entry deleted sucessfully")
    
        elif s == False:
            messagebox.showerror("Error","Username not found")
    
        elif s == True:
            messagebox.showerror("Error","All fields are required")
    except:
        messagebox.showerror("Error","Unknown error")

#=========================================================

def changepass(userid,passward,newpassward,cnewpassward):
    try:
        if newpassward != cnewpassward :
            messagebox.showerror("Error","passward and confirm passward should be same")
            return
    
        s=login(userid,passward)
    
        if s == "admin" :
            #f=open("attendence/adminlogin.dat","rb")
            f=open("adminlogin.dat","rb")
            data=f.readlines()
            f.close()
            f=open("adminlogin.dat","wb")
            for i in data:
                if userid in str(i):
                    continue
                dump(i,f) 
            ne= userid+" "+newpassward+"\n"
            dump(ne,f)  
            f.close()
            messagebox.showinfo("Sucess","Password change sucessfully")
    
        elif s == "login":
            #f=open("attendence/login.dat","rb") 
            f=open("login.dat","rb")
            data=f.readlines()
            f.close()
            f=open("login.dat","wb")
            for i in data:
                if userid in str(i):
                    continue
                dump(i,f)  
            ne=userid+" "+newpassward+"\n"
            dump(ne,f) 
            f.close()
            messagebox.showinfo("Sucess","Password change sucessfully")
    
        elif s == False:
            messagebox.showerror("Error","Username or password not found")
    
        elif s == True:
            messagebox.showerror("Error","All fields are required")
    except:
        messagebox.showerror("Error","Unknown error")

#=========================================================

def createnewws(s,t):
    try:
        wb=load_workbook(s+".xlsx")
        ws1=wb["Sheet1"]
        ws=wb.copy_worksheet(ws1)
        ws.title= t
        f=open("sheet.txt","w")
        f.write(t)
        f.close()
        wb.save(s+".xlsx")
        messagebox.showinfo("Sucess","New sheet created")
    except:
        messagebox.showerror("Error","Unable to create worksheet")

      
#===========================================
def cheackentry(userid):
    try:
        #af=open("attendence/adminlogin.dat","rb")
        #f=open("attendence/login.dat","rb")
        af=open("adminlogin.dat","rb")
        f=open("login.dat","rb")
        for i in f.readlines():
            if userid in str(i) :
                return False
        for i in af.readlines():
            if userid in str(i) :
                return False
        af.close()
        f.close()
        return True
    except:
        messagebox.showerror("Error","Unknown error")

#=======create admin=======
def createadmin(userid,passward,cpassward):
    try:
        if cpassward == passward:
            if cheackentry(userid):
                #f=open("attendence/adminlogin.dat","ab")
                f=open("adminlogin.dat","ab")
                dump(userid+" "+passward+"\n",f)
                f.close()
                messagebox.showinfo("Save","Admin Entry saved")
            else :
                messagebox.showerror("Error","select another username")
        else:
            messagebox.showerror("Error","passward And confirm passward should be same")
    except:
        messagebox.showerror("Error","Unknown error")


#======================= sineup  ==============================================================
   
def sinup(userid,passward,cpassward):
    try:
        if cpassward == passward:
            if cheackentry(userid):
                #f=open("attendence/login.dat","ab")
                f=open("login.dat","ab")
                dump(userid+" "+passward+"\n",f)
                f.close()
                messagebox.showinfo("Save","User Entry saved")
            else :
                messagebox.showerror("Error","select another username")
        else:
            messagebox.showerror("Error","passward And confirm passward should be same")
    except:
        messagebox.showerror("Error","Unknown error")
#===========================================================================

#============================================================================
def login(userid,passward):
    try:
        if userid ==" " or passward==" ":
            return True
        #af=open("attendence/adminlogin.dat","rb")
        #f=open("attendence/login.dat","rb")
        af=open("adminlogin.dat","rb")
        f=open("login.dat","rb")
        for r in af.readlines() :
            if userid in str(r) and passward in str(r) :
                return "admin"
        for r in f.readlines():
            if userid in str(r) and passward in str(r) :
                return "login"
        return False
    except:
        messagebox.showerror("Error","Unknown error")


def test(userid,passward):
    try:
        temp=Frame(page1,bg="azure2")
        temp.place(x=750,y=5,height=150,width=700)
#=====================
        Label(temp,text="Login as : ",font=('Goudy old style',13),bg="azure2",fg="black",bd=0).place(x=500,y=60)
        Label(temp,text= userid,font=('Goudy old style',20),bg="azure2",fg="red",bd=0).place(x=500,y=80)
#=====================
        screen=login(userid,passward)
        if screen =="admin":
            temp=Frame(page1,bg="dodgerblue4")
            temp.place(x=20,y=410,height=700,width=250)
            Button(temp, text='Create Admin Login', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command=lambda: admin()).place(x=30, y=10,width=210,height=40)   
            Button(temp, text='Create User Login', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command=lambda: user() ).place(x=30, y=70,width=200,height=40)
            Button(temp, text='Change passward', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : change() ).place(x=30, y=130,width=200,height=40)
            Button(temp, text='Delete Login', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : delete() ).place(x=30, y=190,width=200,height=40)
            Button(temp, text='Log out', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : log_out()).place(x=35, y=250,width=180,height=40)
            home()
            page1.tkraise()
        elif screen =="login":
            temp=Frame(page1,bg="dodgerblue4")
            temp.place(x=20,y=410,height=700,width=250)
            Button(temp, text='Change passward', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : change() ).place(x=30, y=10,width=200,height=40)
            Button(temp, text='Log out', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : log_out()).place(x=35, y=60,width=180,height=40)
            home()
            page1.tkraise()
        elif screen ==False:
            messagebox.showerror("Error","Username or password not found")
        elif screen==True:
            messagebox.showerror("Error","All fields are required")
    except:
        messagebox.showerror("Error","Unknown error")

#==============================================


def numberofentry(ws):
    n=0
    while True:
        if ws['A'+str(n+10)].value == None  :
            break
        n=n+1
    return n

def cheackplace(ws):
    for row in range(4,23):
        char = get_column_letter(row)
        if ws[char+'9'].value == None  :
            break
    return char

def entertheory(s,date,time,g):
    try:
        try:
            wb=load_workbook(s+'.xlsx')
        except:
            messagebox.showerror("Error","Workbook not found")
        
        if len(wb.sheetnames)==1:
            messagebox.showerror("","Create new sheet")
            create() 
            return
        f=open("sheet.txt","r")
        w=f.readline()
        f.close()
        ws=wb[w]
        char=cheackplace(ws)
        n=numberofentry(ws)
        
        ws[char+'9'].value = date+time
        r=[]
        for col in range(10,n+11):
            r.append(str(ws['B'+str(col)].value))
        a=[int(i) for i in g.split(",")]
        '''a=[]
        for i in g.split(","):
            n=0
            for j in i:
                if j.isalpha():
                    continue
                n=n*10+int(j)
            a.append(n)'''
        for col in range(10,n+10):
            if int(r[col-10][-3::]) in a :
                ws[char+str(col)] =  'P'
            else :
                ws[char+str(col)] =  'A'
        wb.save(s+".xlsx")
        messagebox.showinfo("save","Entry save sucessfully")
        if char=="V":
            messagebox.showerror("","Sheet is full! Create new sheet")
            create()
    except:
        messagebox.showerror("An Error occure","Entry not saved")


def enterpract(s,date,time,g):
    try:
        try:
            wb=load_workbook(s+'.xlsx')
        except:
            messagebox.showerror("Error","Workbook not found")
        
        if len(wb.sheetnames)==1:
            messagebox.showerror("","Create new sheet")
            create()
            return
        f=open("sheet.txt","r")
        w=f.readline()
        f.close()
        ws=wb[w]
        n=numberofentry(ws)
        for row in range(24,32):
            char = get_column_letter(row)
            if ws[char+'9'].value == None  :
                break
        ws[char+'9'].value = date+time
        r=[]
        for col in range(10,n+11):
            r.append(str(ws['B'+str(col)].value))
        a=[int(i) for i in g.split(",")]
        '''a=[]
        for i in g.split(","):
            n=0
            for j in i:
                if j.isalpha():
                    continue
                n=n*10+int(j)
            a.append(n)'''
        for col in range(10,n+10):
            if int(r[col-10][-3::]) in a :
                ws[char+str(col)] =  'P'
            else :
               ws[char+str(col)] =  'A'
        wb.save(s+".xlsx")
        messagebox.showinfo("save","Entry save sucessfully")
        
        if char=="AE":
            messagebox.showerror("","Sheet is full! Create new sheet")
            create()
    except:
        messagebox.showerror("An Error occure","Entry not saved")



#========================== f end ==========================================


#==================== login ===========================


#=========Image==========
#img_bg= PhotoImage(file="attendence/pexels .png")
img_bg= PhotoImage(file="pexels .png")
Label(Login,image=img_bg,bg='white').place(x=0,y=0,relheight=1,relwidth=1)
# ============= Page 1 =========
#Login.configure(bg="#fff")
Frame_login=Frame(Login,bg="azure")
Frame_login.place(x=460,y=200,height=400,width=550)
Label(Frame_login,text=" Login ",font=("Microsoft YaHei UI Light",23,"bold"),fg="#57a1f8",bg="azure").place(x=230,y=10) 
   #======user===================================
def on_enter(e):
    txt_user.delete(0,"end")
def on_leave(e):
    name=txt_user.get()
    if name=="":
        txt_user.insert(0,"Username")
   #login level; 
txt_user=Entry(Frame_login,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
txt_user.place(x=155,y=80)
txt_user.insert(0,"Username")
txt_user.bind('<FocusIn>', on_enter)
txt_user.bind('<FocusOut>', on_leave)
Frame(Frame_login,width=295,height=2,bg="black").place(x=145,y=107)
   #======pass==========================
def on_enter(e):
    txt_pass.delete(0,"end")
def on_leave(e):
    name=txt_pass.get()
    if name=="":
        txt_pass.insert(0,"Password")
    # password level 
txt_pass=Entry(Frame_login,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0,show="*")
txt_pass.place(x=155,y=150)
txt_pass.insert(0,"Password")
txt_pass.bind('<FocusIn>', on_enter)
txt_pass.bind('<FocusOut>', on_leave)
Frame(Frame_login,width=295,height=2,bg="black").place(x=145,y=177)
#=======forgot=============
Button(Frame_login,text=" Forget pasword ?",bg="azure",fg="#d77337",bd=0,font=("times new roman",12),command= lambda : forgot(txt_user.get())).place(x=140,y=185)
#=========login===========
Button(Frame_login,width=39,pady=7,text="Login",bg="#57a1f8",fg="white",border=0,command=lambda: test(txt_user.get(),txt_pass.get())).place(x=160,y=230)   



# ======== main page ===========
page1.configure(bg="azure")
Frame_side=Frame(page1,bg="dodgerblue4")
Frame_side.place(x=20,y=10,height=1520,width=250)
Frame_up=Frame(page1,bg="azure2")
Frame_up.place(x=5,y=5,height=150,width=1550)

#=====Image=====
#img= PhotoImage(file='attendence\logoies.png')
img= PhotoImage(file='logoies.png')
Label(page1,image=img,bg='azure2').place(x=5,y=5)


#button
Button(Frame_side, text='Home', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command= lambda : home()).place(x=35, y=180,width=180,height=40)
Button(Frame_side, text='Create new sheet', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command=lambda: create()).place(x=30, y=240,width=200,height=40)
Button(Frame_side, text='Theory', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command=lambda: theory()).place(x=35, y=300,width=180,height=40)
Button(Frame_side, text='Practicle', font=('Goudy old style', 15,"bold"),bd=0,fg="white",bg="dodgerblue4",command=lambda: practicle()).place(x=35, y=360,width=180,height=40)


def home():
    Frame(page1,bg="azure").place(x=270,y=155,height=1365,width=1280)

def theory():
    home()
    Frame_main=Frame(page1,bg="azure")
    Frame_main.place(x=550,y=170,height=600,width=650)
    #class block
    Label(Frame_main,text="Class",font=('Goudy old style', 15, 'bold'),fg="black",bg="azure",bd=0).place(x=30,y=30)
    pg1_class=Entry(Frame_main,font=("times new roman",15),bg="white")
    pg1_class.place(x=30,y=60,width=90,height=33)

    #================================
    Label(Frame_main,text="Theory",font=('Goudy old style', 20, 'bold'),fg="maroon",bg="azure",bd=0).place(x=420,y=60)
    #=========================
    def on_enter(e):
        pg1_time.delete(0,"end")
    def on_leave(e):
        name=pg1_time.get()
        if name=="":
            pg1_time.insert(0,"HH:MM")
# time block
    Label(Frame_main,text="Time",font=('Goudy old style', 15, 'bold'),bg="azure",fg="black",bd=0).place(x=450,y=150)
    pg1_time=Entry(Frame_main,font=("Microsoft YaHei UI Light",13),bg="White")
    pg1_time.place(x=450,y=175,width=150,height=30)
    pg1_time.insert(0,"HH:MM")
    pg1_time.bind('<FocusIn>', on_enter)
    pg1_time.bind('<FocusOut>', on_leave)

#==================
    def on_enter(e):
        pg1_date.delete(0,"end")
    def on_leave(e):
        name=pg1_date.get()
        if name=="":
            pg1_date.insert(0,"DD/MM/YYYY")
# date block 
    Label(Frame_main,text="Date",font=('Goudy old style', 15, 'bold'),bg="azure",fg="black",bd=0).place(x=30,y=150)
    pg1_date=Entry(Frame_main,font=("Microsoft YaHei UI Light",13),bg="white")
    pg1_date.place(x=30,y=175,width=150,height=30)
    pg1_date.insert(0,"DD/MM/YYYY")
    pg1_date.bind('<FocusIn>', on_enter)
    pg1_date.bind('<FocusOut>', on_leave)

# roll no block
    Label(Frame_main,text="Roll number",font=('Goudy old style', 20, 'bold'),bg="azure",fg="black",bd=0).place(x=30,y=250)
    pg1_roll=Entry(Frame_main,font=("times new roman",15),bg="White")
    pg1_roll.place(x=30,y=285,width=600,height=150)
    
    
    Button(Frame_main, text='Submit', font=('Goudy old style', 15,"bold"),bg="dodgerblue4",fg='white',command=lambda: entertheory(pg1_class.get(),pg1_date.get(),pg1_time.get(),pg1_roll.get())).place(x=290, y=500)




def practicle():
    home()
    Frame_main=Frame(page1,bg="azure")
    Frame_main.place(x=550,y=170,height=600,width=650)
    #class block
    Label(Frame_main,text="Class",font=('Goudy old style', 15, 'bold'),fg="black",bg="azure",bd=0).place(x=30,y=30)
    pg1_class=Entry(Frame_main,font=("times new roman",15),bg="white")
    pg1_class.place(x=30,y=60,width=90,height=33)
    #===============================
    Label(Frame_main,text="Practical",font=('Goudy old style', 20, 'bold'),fg="maroon",bg="azure",bd=0).place(x=420,y=60)
    #=========================
    def on_enter(e):
        pg1_time.delete(0,"end")
    def on_leave(e):
        name=pg1_time.get()
        if name=="":
            pg1_time.insert(0,"HH:MM")
# time block
    Label(Frame_main,text="Time",font=('Goudy old style', 15, 'bold'),bg="azure",fg="black",bd=0).place(x=450,y=150)
    pg1_time=Entry(Frame_main,font=("Microsoft YaHei UI Light",13),bg="White")
    pg1_time.place(x=450,y=175,width=150,height=30)
    pg1_time.insert(0,"HH:MM")
    pg1_time.bind('<FocusIn>', on_enter)
    pg1_time.bind('<FocusOut>', on_leave)

#==================
    def on_enter(e):
        pg1_date.delete(0,"end")
    def on_leave(e):
        name=pg1_date.get()
        if name=="":
            pg1_date.insert(0,"DD/MM/YYYY")
# date block 
    Label(Frame_main,text="Date",font=('Goudy old style', 15, 'bold'),bg="azure",fg="black",bd=0).place(x=30,y=150)
    pg1_date=Entry(Frame_main,font=("Microsoft YaHei UI Light",13),bg="white")
    pg1_date.place(x=30,y=175,width=150,height=30)
    pg1_date.insert(0,"DD/MM/YYYY")
    pg1_date.bind('<FocusIn>', on_enter)
    pg1_date.bind('<FocusOut>', on_leave)

# roll no block
    Label(Frame_main,text="Roll number",font=('Goudy old style', 20, 'bold'),bg="azure",fg="black",bd=0).place(x=30,y=250)
    pg1_roll=Entry(Frame_main,font=("times new roman",15),bg="White")
    pg1_roll.place(x=30,y=285,width=600,height=150)


    Button(Frame_main, text='Submit', font=('Goudy old style', 15,"bold"),bg="dodgerblue4",fg='white',command=lambda: enterpract(pg1_class.get(),pg1_date.get(),pg1_time.get(),pg1_roll.get())).place(x=290, y=500)




def admin(): 
#======= admin sineup =============
    admine=Frame(page1,bg="azure")
    admine.place(x=270,y=155,height=1365,width=1280)
    Label(admine,text="Admin Sign Up",font=("Microsoft YaHei UI Light",21,"bold"),fg="#57a1f8",bg="azure").place(x=290,y=80) 
   #======user===================================
    def on_enter(e):
        tar.delete(0,"end")
    def on_leave(e):
        name=tar.get()
        if name=="":
            tar.insert(0,"Create Username")
   #login level; 
    tar=Entry(admine,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tar.place(x=255,y=170)
    tar.insert(0,"Create Username")
    tar.bind('<FocusIn>', on_enter)
    tar.bind('<FocusOut>', on_leave)
    Frame(admine,width=295,height=2,bg="black").place(x=245,y=197)
   #======pass==========================
    def on_enter(e):
        tacs.delete(0,"end")
    def on_leave(e):
        name=tacs.get()
        if name=="":
            tacs.insert(0,"Enter Password")
    # password level 
    tacs=Entry(admine,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tacs.place(x=255,y=240)
    tacs.insert(0,"Enter Password")
    tacs.bind('<FocusIn>', on_enter)
    tacs.bind('<FocusOut>', on_leave)
    Frame(admine,width=295,height=2,bg="black").place(x=245,y=267)

#=======================
    def on_enter(e):
        tas.delete(0,"end")
    def on_leave(e):
        name=tas.get()
        if name=="":
            tas.insert(0,"Confirm Password")
    tas=Entry(admine,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tas.place(x=255,y=310)
    tas.insert(0,"Confirm Password")
    tas.bind('<FocusIn>', on_enter)
    tas.bind('<FocusOut>', on_leave)
    Frame(admine,width=295,height=2,bg="black").place(x=245,y=337)

#=========  save ===========
    Button(admine,width=10,pady=7,text="Save",font=("times new roman",15),bg="dodgerblue4",fg="white",border=0,command= lambda :createadmin(tar.get(),tacs.get(),tas.get())).place(x=250,y=400)



def user():
#======= user sineup =============
    sineup=Frame(page1,bg="azure")
    sineup.place(x=270,y=155,height=1365,width=1280)
    Label(sineup,text="User Sign Up",font=("Microsoft YaHei UI Light",23,"bold"),fg="#57a1f8",bg="azure").place(x=290,y=80) 
   #======user===================================
    def on_enter(e):
        tr.delete(0,"end")
    def on_leave(e):
        name=tr.get()
        if name=="":
            tr.insert(0,"Create Username")
   #login level; 
    tr=Entry(sineup,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tr.place(x=255,y=170)
    tr.insert(0,"Create Username")
    tr.bind('<FocusIn>', on_enter)
    tr.bind('<FocusOut>', on_leave)
    Frame(sineup,width=295,height=2,bg="black").place(x=245,y=197)
   #======pass==========================
    def on_enter(e):
        tcs.delete(0,"end")
    def on_leave(e):
        name=tcs.get()
        if name=="":
            tcs.insert(0,"Password")
    # password level 
    tcs=Entry(sineup,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tcs.place(x=255,y=240)
    tcs.insert(0,"Enter Password")
    tcs.bind('<FocusIn>', on_enter)
    tcs.bind('<FocusOut>', on_leave)
    Frame(sineup,width=295,height=2,bg="black").place(x=245,y=267)

#=======================
    def on_enter(e):
        ts.delete(0,"end")
    def on_leave(e):
        name=ts.get()
        if name=="":
            ts.insert(0,"Confirm Password")
    ts=Entry(sineup,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    ts.place(x=255,y=310)
    ts.insert(0,"Confirm Password")
    ts.bind('<FocusIn>', on_enter)
    ts.bind('<FocusOut>', on_leave)
    Frame(sineup,width=295,height=2,bg="black").place(x=245,y=337)

#=========  save ===========
    Button(sineup,width=10,pady=7,text="Save",font=("times new roman",15),bg="dodgerblue4",fg="white",border=0,command= lambda :sinup(tr.get(),tcs.get(),ts.get())).place(x=250,y=400)


def create():
    home()
    Frame_main=Frame(page1,bg="azure")
    Frame_main.place(x=550,y=170,height=600,width=650)
    Label(Frame_main,text="Class name",font=('Goudy old style', 15, 'bold'),fg="black",bg="white",bd=0).place(x=30,y=70)
    class_entry=Entry(Frame_main,font=("times new roman",15),bg="white")
    class_entry.place(x=30,y=100,width=100,height=30)

    Label(Frame_main,text="Sheet name",font=('Goudy old style', 20, 'bold'),bg="White",fg="black",bd=0).place(x=30,y=210)
    sheet_entry=Entry(Frame_main,font=("times new roman",15),bg="White")
    sheet_entry.place(x=30,y=245,width=600,height=35)

    Button(Frame_main, text='Save', font=('Goudy old style', 15,"bold"),bg="dodgerblue4",fg='white',command=lambda : createnewws(class_entry.get(),sheet_entry.get())).place(x=100, y=320)

def change():
    #======= change sineup =============
    chang=Frame(page1,bg="azure")
    chang.place(x=270,y=155,height=1365,width=1280)
    Label(chang,text="Change password",font=("Microsoft YaHei UI Light",21,"bold"),fg="#57a1f8",bg="azure").place(x=290,y=80) 
   #======user===================================
    def on_enter(e):
        tar.delete(0,"end")
    def on_leave(e):
        name=tar.get()
        if name=="":
            tar.insert(0,"Username")
   #login level; 
    tar=Entry(chang,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tar.place(x=255,y=170)
    tar.insert(0,"Username")
    tar.bind('<FocusIn>', on_enter)
    tar.bind('<FocusOut>', on_leave)
    Frame(chang,width=295,height=2,bg="black").place(x=245,y=197)
   #======pass==========================
    def on_enter(e):
        tacs.delete(0,"end")
    def on_leave(e):
        name=tacs.get()
        if name=="":
            tacs.insert(0,"Old Password")
    # password level 
    tacs=Entry(chang,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tacs.place(x=255,y=240)
    tacs.insert(0,"Old Password")
    tacs.bind('<FocusIn>', on_enter)
    tacs.bind('<FocusOut>', on_leave)
    Frame(chang,width=295,height=2,bg="black").place(x=245,y=267)

#=======================
    def on_enter(e):
        tas.delete(0,"end")
    def on_leave(e):
        name=tas.get()
        if name=="":
            tas.insert(0,"New Password")
    tas=Entry(chang,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tas.place(x=255,y=310)
    tas.insert(0,"New Password")
    tas.bind('<FocusIn>', on_enter)
    tas.bind('<FocusOut>', on_leave)
    Frame(chang,width=295,height=2,bg="black").place(x=245,y=337)

#=======================
    def on_enter(e):
        t.delete(0,"end")
    def on_leave(e):
        name=t.get()
        if name=="":
            t.insert(0,"Confirm New Password")
    t=Entry(chang,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    t.place(x=255,y=380)
    t.insert(0,"Confirm new Password")
    t.bind('<FocusIn>', on_enter)
    t.bind('<FocusOut>', on_leave)
    Frame(chang,width=295,height=2,bg="black").place(x=245,y=407)

#=========  save ===========
    Button(chang,width=10,pady=7,text="Save",font=("times new roman",15),bg="dodgerblue4",fg="white",border=0,command= lambda :changepass(tar.get(),tacs.get(),tas.get(),t.get())).place(x=250,y=470)


def delete():
    #======= user sineup =============
    sineup=Frame(page1,bg="azure")
    sineup.place(x=270,y=155,height=1365,width=1280)
    Label(sineup,text="Delete account",font=("Microsoft YaHei UI Light",23,"bold"),fg="#57a1f8",bg="azure").place(x=280,y=80) 
   #======user===================================
    def on_enter(e):
        tr.delete(0,"end")
    def on_leave(e):
        name=tr.get()
        if name=="":
            tr.insert(0,"Enter Username")
   #login level; 
    tr=Entry(sineup,font=("Microsoft YaHei UI Light",11),bg="azure",width=25,fg="black",border=0)
    tr.place(x=255,y=170)
    tr.insert(0,"Create Username")
    tr.bind('<FocusIn>', on_enter)
    tr.bind('<FocusOut>', on_leave)
    Frame(sineup,width=295,height=2,bg="black").place(x=245,y=197)

#=========  save ===========
    Button(sineup,width=10,pady=7,text="Confirm",font=("times new roman",15),bg="dodgerblue4",fg="white",border=0,command= lambda :deletentry(tr.get())).place(x=250,y=250)


window.mainloop()

#forget / password 
#add mail of sir 
#create sheet 
#1,2,3,5,6,8,13,24,15,19,26,30,31,32,33,34,38,39,41,43,51,65,55,58,69,135